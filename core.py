# -*- coding: utf-8 -*-
"""
core.py — verification logic for the РСВУ Web of Science indicators,
computed from the Web of Science **Expanded API** (Core Collection times-cited).

Why the API (not the InCites export): the export leaves the Core-Collection
"Times Cited" empty, so only InCites counts (~20% higher) were available. The
Expanded API returns Core-Collection times-cited per record
(dynamic_data.citation_related.tc_list.silo_tc with coll_id="WOS"), which is the
count РСВУ actually used.

Strategy: ONE master fetch of all university records in the window
(OG=... AND PY=...), parse each record, then filter locally by Research Area
(ascatype="extended") into each professional field (ПН) via the matrix, and
recompute:

    L h_index       = Индекс на цитируемост на Хирш (WoS)
    N avg_citations = Среден брой цитирания на документ (WoS)
    P cited_once    = Документи, цитирани поне веднъж (WoS)
    R articles      = Статии в научни списания (WoS)        (Document Type = Article)
    U collaborative = Съвместни научни публикации (WoS)      (>=2 institutions)

The extractor functions are deliberately isolated and defensive. Confirm their
field paths against your subscription with probe.py before trusting numbers.
"""

import io
import re
import time
from collections import OrderedDict
from dataclasses import dataclass


@dataclass
class Config:
    base_url: str = "https://wos-api.clarivate.com/api/wos"
    database: str = "WOS"                 # Core Collection
    edition: str = ""                     # "" = all editions in your subscription
    org_query: str = "OG=(Medical University Varna)"
    py_from: int = 2021
    py_to: int = 2025
    article_doctypes: tuple = ("Article",)
    proceedings_doctypes: tuple = ("Proceedings Paper",)
    count_review_as_article: bool = False
    collab_min_institutions: int = 2      # >=2 distinct institutions => collaborative
    page_size: int = 100                  # Expanded hard max
    sleep_between_calls: float = 1.0
    max_retries: int = 4

    def article_type_set(self):
        s = set(self.article_doctypes)
        if self.count_review_as_article:
            s.add("Review")
        return s


# Информация_ПН_ВУ_*.xlsx layout
INFO_COLS = OrderedDict([
    ("L", ("h_index",        12)),
    ("N", ("avg_citations",  14)),
    ("P", ("cited_once",     16)),
    ("R", ("articles",       18)),
    ("T", ("proceedings",    20)),
    ("U", ("collaborative",  21)),
])
PN_CODE_COL = 3
PN_NAME_COL = 4
METRIC_KEYS = ("h_index", "avg_citations", "cited_once", "articles", "proceedings", "collaborative")

def pn_code_to_matrix(code):
    return "ПН" + str(int(code))


# ----------------------------------------------------------------------------
# Research-area name handling
# ----------------------------------------------------------------------------
SU_CANONICAL = {
    "SCIENCE & TECHNOLOGY OTHER TOPICS": "Science & Technology - Other Topics",
    "LIFE SCIENCES BIOMEDICINE OTHER TOPICS": "Life Sciences & Biomedicine - Other Topics",
    "PHYSICAL SCIENCES OTHER TOPICS": "Physical Sciences - Other Topics",
    "SOCIAL SCIENCES OTHER TOPICS": "Social Sciences - Other Topics",
    "ARTS HUMANITIES OTHER TOPICS": "Arts & Humanities - Other Topics",
}

def canonical_area(name):
    name = " ".join(str(name).split()).strip()
    up = name.upper()
    if up in SU_CANONICAL:
        return SU_CANONICAL[up]
    return name.title().replace(" And ", " & ")

def norm_key(s):
    return re.sub(r"[^A-Z0-9]", "", str(s).upper())


# ----------------------------------------------------------------------------
# Input readers (matrix is now .xlsx; info is .xlsx)
# ----------------------------------------------------------------------------
def load_area_matrix(path_or_buffer):
    """Return (mapping {ПНxxx:[display areas]}, keysets {ПНxxx:set(norm keys)}).
    Layout: row1 = ПН names, row2 = ПН codes, col1 = area name, data from row3.
    Reads .xlsx (openpyxl) or .xls (xlrd) transparently."""
    name = getattr(path_or_buffer, "name", "") or (path_or_buffer if isinstance(path_or_buffer, str) else "")
    if str(name).lower().endswith(".xls"):
        import xlrd
        if hasattr(path_or_buffer, "read"):
            wb = xlrd.open_workbook(file_contents=path_or_buffer.read())
        else:
            wb = xlrd.open_workbook(path_or_buffer)
        sh = wb.sheets()[0]
        get = lambda r, c: sh.cell_value(r - 1, c - 1)
        nrows, ncols = sh.nrows, sh.ncols
    else:
        import openpyxl
        wb = openpyxl.load_workbook(path_or_buffer, data_only=True)
        ws = wb.worksheets[0]
        get = lambda r, c: ws.cell(r, c).value
        nrows, ncols = ws.max_row, ws.max_column

    codes = [get(2, c) for c in range(1, ncols + 1)]
    mapping, keysets = {}, {}
    for c, code in enumerate(codes, start=1):
        code = str(code).strip() if code is not None else ""
        if not code.startswith("ПН"):
            continue
        areas, keys = [], set()
        for r in range(3, nrows + 1):
            raw = get(r, 1)
            if raw and get(r, c) in (1, 1.0, "1"):
                areas.append(canonical_area(raw))
                keys.add(norm_key(raw))
        if areas:
            mapping[code] = areas
            keysets[code] = keys
    return mapping, keysets


def load_info(path_or_buffer):
    import openpyxl
    wb = openpyxl.load_workbook(path_or_buffer, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
    rows = []
    for r in range(2, ws.max_row + 1):
        code = ws.cell(r, PN_CODE_COL).value
        if code in (None, ""):
            continue
        rec = {"pn_code": int(code),
               "pn_name": ws.cell(r, PN_NAME_COL).value, "expected": {}}
        for _, (key, col_idx) in INFO_COLS.items():
            rec["expected"][key] = ws.cell(r, col_idx).value
        rows.append(rec)
    return rows


def build_vocab(keysets):
    return sorted({k for s in keysets.values() for k in s}, key=len, reverse=True)


# ----------------------------------------------------------------------------
# Queries
# ----------------------------------------------------------------------------
def master_query(cfg):
    return "{org} AND PY=({a}-{b})".format(org=cfg.org_query, a=cfg.py_from, b=cfg.py_to)


# ----------------------------------------------------------------------------
# Expanded REC extractors  (CONFIRM against probe.py output)
# ----------------------------------------------------------------------------
def extract_uid(rec):
    return rec.get("UID") or rec.get("uid")

def extract_times_cited(rec):
    """Core-Collection times cited: silo_tc with coll_id='WOS'. Falls back to max."""
    try:
        silo = rec["dynamic_data"]["citation_related"]["tc_list"]["silo_tc"]
        if isinstance(silo, dict):
            silo = [silo]
        wos = [int(s.get("local_count", 0) or 0) for s in silo if s.get("coll_id") == "WOS"]
        if wos:
            return max(wos)
        allc = [int(s.get("local_count", 0) or 0) for s in silo]
        return max(allc) if allc else 0
    except Exception:
        return 0

def extract_doctypes(rec):
    try:
        dt = rec["static_data"]["summary"]["doctypes"]["doctype"]
        if isinstance(dt, str):
            return [dt]
        if isinstance(dt, list):
            return [str(x) for x in dt]
        return [str(dt)]
    except Exception:
        return []

def extract_research_area_keys(rec):
    """Normalized keys of the record's WoS Research Areas (ascatype='extended')."""
    keys = set()
    try:
        subj = rec["static_data"]["fullrecord_metadata"]["category_info"]["subjects"]["subject"]
        if isinstance(subj, dict):
            subj = [subj]
        ext = [s for s in subj if isinstance(s, dict) and s.get("ascatype") == "extended"]
        for s in ext:
            val = s.get("content", "")
            if val:
                keys.add(norm_key(val))
    except Exception:
        pass
    return keys

def extract_num_institutions(rec):
    """Distinct institutions across addresses (one preferred org per address)."""
    orgs = set()
    try:
        addrs = rec["static_data"]["fullrecord_metadata"]["addresses"]
        alist = addrs.get("address_name", [])
        if isinstance(alist, dict):
            alist = [alist]
        for a in alist:
            ol = a.get("address_spec", {}).get("organizations", {}).get("organization", [])
            if isinstance(ol, dict):
                ol = [ol]
            picked = None
            for entry in ol:
                if isinstance(entry, dict):
                    val = entry.get("content", "")
                    if entry.get("pref") == "Y":
                        picked = val
                    elif picked is None:
                        picked = val
                else:
                    picked = str(entry)
            if picked:
                orgs.add(picked.strip().lower())
    except Exception:
        pass
    return len(orgs)

def probe_record(rec):
    """Human-readable view of what the extractors pull from one raw REC."""
    return {
        "UID": extract_uid(rec),
        "times_cited(WOS)": extract_times_cited(rec),
        "doctypes": extract_doctypes(rec),
        "research_areas(extended,norm)": sorted(extract_research_area_keys(rec)),
        "num_institutions": extract_num_institutions(rec),
    }


# ----------------------------------------------------------------------------
# API access
# ----------------------------------------------------------------------------
def _dig_total(data):
    try:
        return int(data["QueryResult"]["RecordsFound"])
    except Exception:
        return None

def _dig_records(data):
    try:
        recs = data["Data"]["Records"]["records"]["REC"]
        return [recs] if isinstance(recs, dict) else recs
    except Exception:
        return []

def api_get(cfg, api_key, usr_query, first_record, count):
    import requests
    params = {"databaseId": cfg.database, "usrQuery": usr_query,
              "count": count, "firstRecord": first_record}
    if cfg.edition:
        params["edition"] = cfg.edition
    headers = {"X-ApiKey": api_key, "Accept": "application/json"}
    last = None
    for attempt in range(cfg.max_retries):
        try:
            resp = requests.get(cfg.base_url, headers=headers, params=params, timeout=60)
            if resp.status_code == 429:
                time.sleep((attempt + 1) * 5); continue
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            last = e
            time.sleep((attempt + 1) * 2)
    raise RuntimeError("WoS API call failed after retries: %s" % last)

def fetch_all(cfg, api_key, usr_query, progress=None):
    first = api_get(cfg, api_key, usr_query, 1, cfg.page_size)
    total = _dig_total(first)
    recs = _dig_records(first)
    if total is None:
        total = len(recs)
    got = len(recs)
    if progress:
        progress(got, total)
    while got < total:
        time.sleep(cfg.sleep_between_calls)
        page = _dig_records(api_get(cfg, api_key, usr_query, got + 1, cfg.page_size))
        if not page:
            break
        recs.extend(page)
        got += len(page)
        if progress:
            progress(got, total)
    return recs, total

def fetch_sample(cfg, api_key, usr_query, n=2):
    """Fetch up to n raw records for schema inspection."""
    data = api_get(cfg, api_key, usr_query, 1, max(1, min(n, cfg.page_size)))
    return _dig_records(data), _dig_total(data)


# ----------------------------------------------------------------------------
# Parse, filter, metrics
# ----------------------------------------------------------------------------
def parse_records(records):
    seen, out = set(), []
    for rec in records:
        uid = extract_uid(rec)
        if uid in seen:
            continue
        seen.add(uid)
        out.append({
            "uid": uid,
            "tc": extract_times_cited(rec),
            "doctypes": extract_doctypes(rec),
            "ninst": extract_num_institutions(rec),
            "area_keys": extract_research_area_keys(rec),
        })
    return out

def filter_by_areas(parsed, keyset):
    return [d for d in parsed if d["area_keys"] & keyset]

def h_index(cites):
    h = 0
    for i, c in enumerate(sorted(cites, reverse=True), start=1):
        if c >= i:
            h = i
        else:
            break
    return h

def compute_metrics(parsed, cfg):
    tcs = [d["tc"] for d in parsed]
    n = len(tcs)
    art = cfg.article_type_set()
    proc = set(cfg.proceedings_doctypes)
    return {
        "n_documents": n,
        "h_index": h_index(tcs),
        "avg_citations": round(sum(tcs) / n, 2) if n else 0,
        "cited_once": sum(1 for t in tcs if t >= 1),
        "articles": sum(1 for d in parsed if any(x in art for x in d["doctypes"])),
        "proceedings": sum(1 for d in parsed if any(x in proc for x in d["doctypes"])),
        "collaborative": sum(1 for d in parsed if d["ninst"] >= cfg.collab_min_institutions),
    }

def compare(expected, computed):
    out = OrderedDict()
    for key in METRIC_KEYS:
        exp, got = expected.get(key), computed.get(key)
        try:
            if key == "avg_citations":
                match = exp is not None and abs(float(exp) - float(got)) <= 0.05
            else:
                match = exp is not None and int(exp) == int(got)
        except Exception:
            match = False
        out[key] = {"expected": exp, "computed": got, "match": match}
    return out

def verify_all(parsed, info, keysets, cfg):
    results = []
    for r in info:
        mcode = pn_code_to_matrix(r["pn_code"])
        keyset = keysets.get(mcode)
        if not keyset:
            results.append({"pn": r["pn_code"], "name": r["pn_name"],
                            "n_documents": None, "cmp": None, "note": "no areas mapped (n/a)"})
            continue
        sub = filter_by_areas(parsed, keyset)
        computed = compute_metrics(sub, cfg)
        results.append({"pn": r["pn_code"], "name": r["pn_name"],
                        "n_documents": computed["n_documents"],
                        "cmp": compare(r["expected"], computed), "note": ""})
    return results


def report_to_xlsx_bytes(results, cfg=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WoS verification"
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    if cfg:
        ws.append(["Settings:", "org=%s" % cfg.org_query,
                   "years=%d-%d" % (cfg.py_from, cfg.py_to),
                   "collab>=%d inst" % cfg.collab_min_institutions,
                   "reviews_as_articles=%s" % cfg.count_review_as_article]); ws.append([])
    hdr = ["ПН", "Направление", "Docs"]
    for key in METRIC_KEYS:
        hdr += [key + " (file)", key + " (calc)", key + " match"]
    ws.append(hdr)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for r in results:
        if not r["cmp"]:
            ws.append([r["pn"], r["name"], r.get("note", "n/a")]); continue
        row = [r["pn"], r["name"], r["n_documents"]]
        for key in METRIC_KEYS:
            v = r["cmp"][key]
            row += [v["expected"], v["computed"], "YES" if v["match"] else "NO"]
        ws.append(row)
        last, col = ws.max_row, 4
        for key in METRIC_KEYS:
            ws.cell(last, col + 2).fill = green if r["cmp"][key]["match"] else red
            col += 3
    for col_cells in ws.columns:
        w = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(w, 45)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


# ----------------------------------------------------------------------------
# Offline self-test (extractors + metrics on a synthetic record)
# ----------------------------------------------------------------------------
def _selftest():
    cfg = Config()
    rec = {
        "UID": "WOS:000111",
        "dynamic_data": {"citation_related": {"tc_list": {"silo_tc": [
            {"coll_id": "WOS", "local_count": 7}, {"coll_id": "BCI", "local_count": 9}]}}},
        "static_data": {
            "summary": {"doctypes": {"doctype": "Article"}},
            "fullrecord_metadata": {
                "category_info": {"subjects": {"subject": [
                    {"ascatype": "traditional", "code": "PY", "content": "Medicine, General & Internal"},
                    {"ascatype": "extended", "content": "General & Internal Medicine"}]}},
                "addresses": {"address_name": [
                    {"address_spec": {"organizations": {"organization": [
                        {"pref": "N", "content": "Med Univ Varna"}, {"pref": "Y", "content": "Medical University Varna"}]}}},
                    {"address_spec": {"organizations": {"organization": {"pref": "Y", "content": "Sofia University"}}}},
                ]},
            },
        },
    }
    p = probe_record(rec)
    assert p["times_cited(WOS)"] == 7, p          # WOS silo, not BCI=9
    assert p["doctypes"] == ["Article"], p
    assert p["research_areas(extended,norm)"] == [norm_key("General & Internal Medicine")], p
    assert p["num_institutions"] == 2, p
    # second record: a Proceedings Paper, single institution, 0 cites
    rec2 = {
        "UID": "WOS:000222",
        "dynamic_data": {"citation_related": {"tc_list": {"silo_tc": {"coll_id": "WOS", "local_count": 0}}}},
        "static_data": {
            "summary": {"doctypes": {"doctype": "Proceedings Paper"}},
            "fullrecord_metadata": {
                "category_info": {"subjects": {"subject": {"ascatype": "extended", "content": "General & Internal Medicine"}}},
                "addresses": {"address_name": {"address_spec": {"organizations": {"organization": {"pref": "Y", "content": "Medical University Varna"}}}}},
            },
        },
    }
    parsed = parse_records([rec, rec, rec2])        # rec deduped; rec2 distinct
    assert len(parsed) == 2
    m = compute_metrics(parsed, cfg)
    assert m == {"n_documents": 2, "h_index": 1, "avg_citations": 3.5,
                 "cited_once": 1, "articles": 1, "proceedings": 1, "collaborative": 1}, m
    return "OK", p, m

if __name__ == "__main__":
    print("core self-test:", _selftest())
